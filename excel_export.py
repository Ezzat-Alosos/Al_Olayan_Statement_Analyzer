"""
محرك تصدير Excel مع الرسوم البيانية - نسخة محسنة بالكامل
يدعم العربية بشكل صحيح مع اتجاه RTL في المخططات
"""

from io import BytesIO
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import time
from typing import Dict, List, Optional
from datetime import datetime
import os

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# استيراد مكتبات معالجة النص العربي
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
try:
    import arabic_reshaper
    from bidi.algorithm import get_display
    HAS_ARABIC = True
except ImportError:
    HAS_ARABIC = False
    arabic_reshaper = None
    get_display = None


def ar(text) -> str:
    """معالجة النص العربي للعرض الصحيح في Matplotlib"""
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    
    if HAS_ARABIC:
        try:
            reshaped = arabic_reshaper.reshape(text)
            bidi_text = get_display(reshaped)
            return bidi_text
        except Exception:
            return text
    return text


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# إعداد الخط العربي لـ Matplotlib
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _setup_arabic_font():
    """محاولة تحميل خط عربي لـ Matplotlib"""
    import platform
    from matplotlib import font_manager
    
    font_candidates = []
    
    # مسارات حسب نظام التشغيل
    if platform.system() == "Windows":
        font_candidates = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/tahoma.ttf",
            "C:/Windows/Fonts/segoeui.ttf",
            "C:/Windows/Fonts/calibri.ttf",
        ]
    elif platform.system() == "Darwin":
        font_candidates = [
            "/System/Library/Fonts/Arial.ttf",
            "/Library/Fonts/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Arial.ttf",
            "/System/Library/Fonts/Supplemental/Tahoma.ttf",
        ]
    else:
        font_candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ]
    
    # مسارات محلية
    project_dir = os.path.dirname(os.path.abspath(__file__))
    font_candidates.extend([
        os.path.join(project_dir, 'fonts', 'Tajawal-Regular.ttf'),
        os.path.join(project_dir, 'fonts', 'Cairo-Regular.ttf'),
        os.path.join(project_dir, 'arial.ttf'),
        os.path.join(project_dir, 'tahoma.ttf'),
    ])
    
    for font_path in font_candidates:
        if os.path.exists(font_path):
            try:
                font_manager.fontManager.addfont(font_path)
                font_name = font_manager.FontProperties(fname=font_path).get_name()
                plt.rcParams['font.family'] = font_name
                plt.rcParams['axes.unicode_minus'] = False  # لعرض علامة السالب بشكل صحيح
                print(f"✅ تم تحميل الخط العربي: {font_path} ({font_name})")
                return font_name
            except Exception as e:
                print(f"⚠️ فشل تحميل الخط {font_path}: {e}")
                continue
    
    print("⚠️ لم يتم العثور على خط عربي")
    return None

_setup_arabic_font()


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# أنماط Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
HEADER_FONT = Font(color="FFFFFF", bold=True, name='Tajawal')
NORMAL_FONT = Font(name='Tajawal', size=11)
BOLD_FONT = Font(name='Tajawal', size=11, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="E0E4E8"),
    right=Side(style="thin", color="E0E4E8"),
    top=Side(style="thin", color="E0E4E8"),
    bottom=Side(style="thin", color="E0E4E8"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize_and_style(worksheet):
    """تنسيق ورقة العمل"""
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A2"
    
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = NORMAL_FONT
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
    
    for column_cells in worksheet.columns:
        max_length = 12
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, min(len(str(cell.value)) + 2, 45))
        worksheet.column_dimensions[column_letter].width = max_length


def _write_sheet(writer, sheet_name: str, frame: pd.DataFrame):
    """كتابة DataFrame في ورقة"""
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    _autosize_and_style(writer.book[sheet_name])


def _shorten_text(text, max_len=10):
    """تقصير النص"""
    text = str(text)
    return text[:max_len] + ".." if len(text) > max_len else text


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# دوال إنشاء الرسوم البيانية - كلها تستخدم ar() للنص العربي
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _create_bar_chart(data, x_col, y_col, title) -> Optional[BytesIO]:
    """إنشاء مخطط شريطي"""
    try:
        fig, ax = plt.subplots(figsize=(9, 5.5), dpi=300)
        x = [ar(_shorten_text(t, 12)) for t in data[x_col].astype(str).tolist()]
        y = data[y_col].tolist()
        colors = plt.cm.Blues(np.linspace(0.4, 0.9, len(x)))[::-1]
        
        bars = ax.bar(x, y, color=colors, edgecolor='darkblue', linewidth=0.5)
        max_y = max(y) if y else 1
        
        for bar, val in zip(bars, y):
            ax.text(bar.get_x() + bar.get_width()/2, 
                    bar.get_height() + max_y * 0.02, 
                    f'{val:,.0f}', 
                    ha='center', va='bottom', fontsize=8, fontweight='bold')
        
        ax.set_ylabel(ar('القيمة'), fontsize=10, fontweight='bold')
        ax.set_title(ar(title), fontsize=12, fontweight='bold', pad=15)
        plt.xticks(rotation=45, ha='right', fontsize=8)
        ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format(int(x), ',')))
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        img_buffer.seek(0)
        return img_buffer
    except Exception as e:
        print(f"⚠️ خطأ في المخطط الشريطي: {e}")
        plt.close()
        return None


def _create_pie_chart(data, x_col, y_col, title) -> Optional[BytesIO]:
    """إنشاء مخطط دائري"""
    try:
        fig, ax = plt.subplots(figsize=(7, 6), dpi=300)
        data = data.sort_values(y_col, ascending=False)
        labels = [ar(_shorten_text(t, 12)) for t in data[x_col].astype(str).tolist()]
        values = data[y_col].tolist()
        
        if sum(values) > 0:
            colors = plt.cm.Blues(np.linspace(0.3, 0.9, len(labels)))[::-1]
            wedges, texts, autotexts = ax.pie(
                values, 
                labels=labels, 
                autopct=lambda p: f'{p:.1f}%' if p > 3 else '', 
                colors=colors, 
                startangle=90, 
                wedgeprops={'edgecolor': 'white', 'linewidth': 1.5}
            )
            ax.set_title(ar(title), fontsize=12, fontweight='bold', pad=15)
            for text in texts:
                text.set_fontsize(9)
            for autotext in autotexts:
                autotext.set_fontsize(9)
                autotext.set_fontweight('bold')
                autotext.set_color('white')
            ax.set_aspect('equal')
        
        plt.tight_layout()
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        img_buffer.seek(0)
        return img_buffer
    except Exception as e:
        print(f"⚠️ خطأ في المخطط الدائري: {e}")
        plt.close()
        return None


def _create_bar_comparison_chart(data, title) -> Optional[BytesIO]:
    """مخطط مقارنة للسنوات"""
    try:
        fig, ax = plt.subplots(figsize=(10, 6), dpi=300)
        years = data.columns[1:].tolist()
        labels = [ar(t) for t in data.iloc[:, 0].tolist()]
        x = np.arange(len(labels))
        width = 0.35
        
        for i, year in enumerate(years):
            values = data[year].tolist()
            offset = width * i - (len(years)-1) * width / 2
            bars = ax.bar(x + offset, values, width, 
                         label=ar(str(year)), 
                         color=plt.cm.Blues(np.linspace(0.4, 0.9, len(years)))[i])
            for bar, val in zip(bars, values):
                ax.text(bar.get_x() + bar.get_width()/2, 
                        bar.get_height() + max(values)*0.02, 
                        f'{val:,.0f}', ha='center', va='bottom', fontsize=7)
        
        ax.set_ylabel(ar('القيمة'), fontsize=11, fontweight='bold')
        ax.set_title(ar(title), fontsize=13, fontweight='bold', pad=15)
        ax.set_xticks(x)
        ax.set_xticklabels(labels, rotation=45, ha='right', fontsize=9)
        ax.legend()
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        plt.tight_layout()
        
        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()
        img_buffer.seek(0)
        return img_buffer
    except Exception as e:
        print(f"⚠️ خطأ في مخطط المقارنة: {e}")
        plt.close()
        return None


def _add_image_to_sheet(worksheet, img_bytes, cell_position: str, 
                        width: int = 380, height: int = 250) -> bool:
    """إدراج صورة في ورقة Excel"""
    if img_bytes is None:
        return False
    
    try:
        if isinstance(img_bytes, bytes):
            img_bytes = BytesIO(img_bytes)
        
        img_bytes.seek(0)
        img = XLImage(img_bytes)
        img.width = width
        img.height = height
        worksheet.add_image(img, cell_position)
        return True
    except Exception as e:
        print(f"⚠️ خطأ في إدراج الصورة: {e}")
        return False


def _write_sheet_with_charts(writer, sheet_name: str, frame: pd.DataFrame, 
                             chart_title: str = None):
    """كتابة ورقة مع المخططات"""
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.book[sheet_name]
    _autosize_and_style(worksheet)
    
    if not frame.empty and len(frame) > 1 and chart_title:
        try:
            chart_start_col = len(frame.columns) + 3
            worksheet.column_dimensions[get_column_letter(len(frame.columns) + 1)].width = 3
            
            # مخطط شريطي
            img_bytes = _create_bar_chart(frame, frame.columns[0], frame.columns[1], 
                                          f"{chart_title} - المخطط الشريطي")
            if img_bytes:
                _add_image_to_sheet(worksheet, img_bytes, 
                                   f"{get_column_letter(chart_start_col)}2", 
                                   width=400, height=260)
            
            # مخطط دائري
            img_bytes_pie = _create_pie_chart(frame, frame.columns[0], frame.columns[1], 
                                              f"{chart_title} - المخطط الدائري")
            if img_bytes_pie:
                _add_image_to_sheet(worksheet, img_bytes_pie, 
                                   f"{get_column_letter(chart_start_col)}18", 
                                   width=380, height=280)
            
            for col in range(chart_start_col, chart_start_col + 3):
                worksheet.column_dimensions[get_column_letter(col)].width = 42
        
        except Exception as e:
            print(f"⚠️ خطأ في إنشاء المخططات لـ {sheet_name}: {e}")


def _write_analysis_sheet(writer, sheet_name: str, frame: pd.DataFrame, 
                          chart_title: str = None):
    """كتابة ورقة تحليل مع مخطط مقارنة"""
    safe_frame = frame.copy()
    if safe_frame.empty:
        safe_frame = pd.DataFrame({"البيان": ["لا توجد بيانات"]})
    safe_frame.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.book[sheet_name]
    _autosize_and_style(worksheet)
    
    if not frame.empty and len(frame) > 1 and chart_title:
        try:
            chart_start_col = len(frame.columns) + 3
            worksheet.column_dimensions[get_column_letter(len(frame.columns) + 1)].width = 3
            
            img_bytes = _create_bar_comparison_chart(frame, f"{chart_title} - مقارنة السنوات")
            if img_bytes:
                _add_image_to_sheet(worksheet, img_bytes, 
                                   f"{get_column_letter(chart_start_col)}2", 
                                   width=450, height=280)
            
            for col in range(chart_start_col, chart_start_col + 3):
                worksheet.column_dimensions[get_column_letter(col)].width = 45
        
        except Exception as e:
            print(f"⚠️ خطأ في إنشاء مخطط التحليل لـ {sheet_name}: {e}")


def generate_excel_report(
    analysis_data: Dict,
    mapped_data: Dict[str, List[float]],
    periods: List[str],
    company_name: str = "",
) -> Optional[BytesIO]:
    """توليد تقرير Excel كامل"""
    start = time.time()
    output = BytesIO()
    
    clean_periods = [str(p).replace("سنة_", "") for p in periods]
    
    # 1. Dashboard
    kpi_data = analysis_data.get("kpi_data", {})
    dashboard_rows = []
    for label, value in kpi_data.items():
        dashboard_rows.append({"المؤشر": label, "القيمة": value if isinstance(value, (int, float)) else 0})
    
    dashboard_df = pd.DataFrame(dashboard_rows)
    health = analysis_data.get("health_score", {})
    if health:
        score_row = {"المؤشر": "درجة الصحة المالية", "القيمة": health.get("total_score", 0)}
        dashboard_df = pd.concat([dashboard_df, pd.DataFrame([score_row])], ignore_index=True)
    
    # 2. القوائم المالية
    bs_keys = ["cash", "receivables", "inventory", "fixed_assets", "total_assets"]
    bs_labels = ["النقدية", "الذمم المدينة", "المخزون", "الأصول الثابتة", "إجمالي الأصول"]
    bs_rows = []
    for key in bs_keys:
        if key in mapped_data:
            bs_rows.append([bs_labels[bs_keys.index(key)]] + mapped_data[key])
    bs_df = pd.DataFrame(bs_rows, columns=["البند"] + clean_periods) if bs_rows else pd.DataFrame()
    
    is_keys = ["revenue", "cost_of_sales", "gross_profit", "operating_income", "net_income"]
    is_labels = ["الإيرادات", "تكلفة المبيعات", "مجمل الربح", "الربح التشغيلي", "صافي الربح"]
    is_rows = []
    for key in is_keys:
        if key in mapped_data:
            is_rows.append([is_labels[is_keys.index(key)]] + mapped_data[key])
    is_df = pd.DataFrame(is_rows, columns=["البند"] + clean_periods) if is_rows else pd.DataFrame()
    
    cf_keys = ["operating_cash_flow", "investing_cash_flow", "financing_cash_flow"]
    cf_labels = ["التدفقات التشغيلية", "التدفقات الاستثمارية", "التدفقات التمويلية"]
    cf_rows = []
    for key in cf_keys:
        if key in mapped_data:
            cf_rows.append([cf_labels[cf_keys.index(key)]] + mapped_data[key])
    cf_df = pd.DataFrame(cf_rows, columns=["البند"] + clean_periods) if cf_rows else pd.DataFrame()
    
    # 3. النسب المالية
    ratios = analysis_data.get("financial_ratios", {})
    ratio_rows = []
    for category in ["liquidity", "profitability", "activity", "leverage"]:
        for key, data in ratios.get(category, {}).items():
            value = data.get("value")
            if value is not None:
                formatted = f"{value:.2f}" if not data.get("is_percentage") else f"{value*100:.2f}%"
                ratio_rows.append([data.get("label", key), formatted, ""])
    ratios_df = pd.DataFrame(ratio_rows, columns=["النسبة", "القيمة", "الحالة"])
    
    # 4. البيانات الخام
    raw_labels = {
        "cash": "النقدية", "receivables": "الذمم المدينة", "inventory": "المخزون",
        "total_assets": "إجمالي الأصول", "total_liabilities": "إجمالي الخصوم",
        "total_equity": "حقوق الملكية", "revenue": "الإيرادات",
        "cost_of_sales": "تكلفة المبيعات", "net_income": "صافي الربح",
        "operating_cash_flow": "التدفقات التشغيلية",
        "investing_cash_flow": "التدفقات الاستثمارية",
        "financing_cash_flow": "التدفقات التمويلية",
    }
    raw_rows = []
    for key, values in mapped_data.items():
        label = raw_labels.get(key, key)
        raw_rows.append([label] + values)
    raw_df = pd.DataFrame(raw_rows, columns=["البند"] + clean_periods)
    
    # 5. كتابة الأوراق
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            _write_sheet(writer, "Dashboard", dashboard_df)
            _write_analysis_sheet(writer, "Balance Sheet", bs_df, "الميزانية العمومية")
            _write_analysis_sheet(writer, "Income Statement", is_df, "قائمة الدخل")
            _write_analysis_sheet(writer, "Cash Flow", cf_df, "التدفقات النقدية")
            _write_sheet(writer, "Financial Ratios", ratios_df)
            _write_sheet(writer, "Raw Data", raw_df)
            
            for worksheet in writer.book.worksheets:
                worksheet.sheet_properties.pageSetUpPr.fitToPage = True
                worksheet.page_setup.fitToWidth = 1
        
        output.seek(0)
        elapsed = time.time() - start
        print(f"⏱️ [excel_export] تم إنشاء الملف في {elapsed:.2f} ثانية")
        return output
    
    except Exception as e:
        print(f"❌ خطأ في إنشاء ملف Excel: {e}")
        import traceback
        traceback.print_exc()
        return None


def get_excel_filename(company_name: str, periods: List[str]) -> str:
    """توليد اسم ملف Excel"""
    date_str = datetime.now().strftime("%Y%m%d")
    period_str = "-".join([str(p) for p in (periods or [])])
    company = (company_name or "تقرير").replace(" ", "_")
    return f"{company}_{period_str}_تحليل_مالي_{date_str}.xlsx"